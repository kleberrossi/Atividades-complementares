# coding: utf-8
import os
import sqlite3
import json
import datetime
import hashlib  # Substituindo bcrypt por hashlib
import base64   # Para codificação/decodificação de salt
import secrets  # Para geração segura de salt
import openpyxl
import logging
import traceback
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, jsonify, g
from werkzeug.utils import secure_filename
from unidecode import unidecode

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session or session["user_type"] != "admin":
            flash("Acesso não autorizado.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

# Função auxiliar para normalizar cabeçalhos/nomes
def normalize_header(text):
    if not isinstance(text, str):
        text = str(text)
    text = unidecode(text)
    return " ".join(text.lower().split())

# Funções para substituir bcrypt com hashlib puro Python
def hash_password(password):
    """Cria um hash seguro para a senha usando hashlib (SHA-256) com salt"""
    # Gerar um salt aleatório
    salt = secrets.token_bytes(16)
    # Criar hash com salt
    h = hashlib.sha256()
    h.update(salt + password.encode("utf-8"))
    password_hash = h.digest()
    # Retornar salt e hash codificados em base64 para armazenamento
    return base64.b64encode(salt).decode("utf-8") + "$" + base64.b64encode(password_hash).decode("utf-8")

def check_password(stored_password, provided_password):
    """Verifica se a senha fornecida corresponde ao hash armazenado"""
    try:
        # Separar salt e hash
        salt_b64, hash_b64 = stored_password.split("$")
        # Decodificar salt
        salt = base64.b64decode(salt_b64)
        # Recriar hash com a senha fornecida
        h = hashlib.sha256()
        h.update(salt + provided_password.encode("utf-8"))
        calculated_hash = h.digest()
        # Comparar com o hash armazenado
        stored_hash = base64.b64decode(hash_b64)
        return secrets.compare_digest(calculated_hash, stored_hash)
    except Exception as e:
        logging.error(f"Erro ao verificar senha: {e}")
        return False

# Configuração de logging
logging.basicConfig(
    filename="app.log",
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = "ej_atividades_complementares_2025"
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

DATABASE = "database.db"



def get_db_connection():
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db_connection(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    conn = get_db_connection()
    conn.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        email TEXT UNIQUE NOT NULL,
        senha TEXT NOT NULL,
        tipo TEXT NOT NULL CHECK(tipo IN ("admin", "aluno"))
    );
    """)
    conn.execute("""
    CREATE TABLE IF NOT EXISTS alunos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER UNIQUE,
        nome TEXT NOT NULL,
        matricula TEXT UNIQUE NOT NULL,
        email TEXT UNIQUE,
        turma TEXT,
        status TEXT DEFAULT 'Ativo' CHECK(status IN ('Ativo', 'Inativo')),
        FOREIGN KEY (usuario_id) REFERENCES usuarios (id)
    );
    """)
    conn.execute("""
    CREATE TABLE IF NOT EXISTS atividades (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        grupo TEXT NOT NULL,
        nome TEXT NOT NULL UNIQUE,
        limite_horas INTEGER,
        tipo_atividade TEXT NOT NULL DEFAULT 'Acadêmica Complementar' CHECK(tipo_atividade IN ('Acadêmica Complementar', 'Extensão Universitária')),
        tem_limitacao BOOLEAN DEFAULT 0,
        tipo_limitacao TEXT CHECK(tipo_limitacao IN ('total', 'semestral')),
        limite_horas_total INTEGER,
        limite_horas_semestral INTEGER
    );
    """)
    conn.execute("""
    CREATE TABLE IF NOT EXISTS requisicoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        aluno_id INTEGER,
        atividade_id INTEGER NOT NULL,
        data_solicitacao TEXT NOT NULL,
        data_evento TEXT NOT NULL,
        horas_solicitadas REAL NOT NULL,
        status TEXT NOT NULL CHECK(status IN ('Pendente', 'Deferida', 'Deferida Parcialmente', 'Indeferida')),
        horas_deferidas REAL,
        observacao TEXT,
        arquivo_comprovante TEXT,
        data_processamento TEXT,
        admin_id INTEGER,
        FOREIGN KEY (aluno_id) REFERENCES usuarios (id),
        FOREIGN KEY (atividade_id) REFERENCES atividades (id),
        FOREIGN KEY (admin_id) REFERENCES usuarios (id)
    );
    """)
    admin_exists = conn.execute("SELECT 1 FROM usuarios WHERE email = ?", ("admin@ej.edu.br",)).fetchone()
    if not admin_exists:
        # Usar novo método de hash para a senha do admin
        hashed_password = hash_password("admin123")
        conn.execute("INSERT INTO usuarios (nome, email, senha, tipo) VALUES (?, ?, ?, ?)",
                     ("Administrador", "admin@ej.edu.br", hashed_password, "admin"))
    
    # Migração: Adicionar coluna tipo_atividade se não existir
    try:
        conn.execute("ALTER TABLE atividades ADD COLUMN tipo_atividade TEXT NOT NULL DEFAULT 'Acadêmica Complementar' CHECK(tipo_atividade IN ('Acadêmica Complementar', 'Extensão Universitária'))")
    except sqlite3.OperationalError:
        # Coluna já existe, ignorar erro
        pass
    
    # Migração: Adicionar colunas de limitação se não existirem
    try:
        conn.execute("ALTER TABLE atividades ADD COLUMN tem_limitacao BOOLEAN DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    
    try:
        conn.execute("ALTER TABLE atividades ADD COLUMN tipo_limitacao TEXT CHECK(tipo_limitacao IN ('total', 'semestral'))")
    except sqlite3.OperationalError:
        pass
    
    try:
        conn.execute("ALTER TABLE atividades ADD COLUMN limite_horas_total INTEGER")
    except sqlite3.OperationalError:
        pass
    
    try:
        conn.execute("ALTER TABLE atividades ADD COLUMN limite_horas_semestral INTEGER")
    except sqlite3.OperationalError:
        pass
    
    # Atualizar atividades existentes que não têm tipo definido
    conn.execute("UPDATE atividades SET tipo_atividade = 'Acadêmica Complementar' WHERE tipo_atividade IS NULL OR tipo_atividade = ''")
    
    conn.commit()

# --- Rotas Admin --- 

@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    conn = get_db_connection()
    total_alunos = conn.execute("SELECT COUNT(*) FROM alunos").fetchone()[0]
    
    # Estatísticas por tipo de atividade
    total_atividades_academicas = conn.execute("SELECT COUNT(*) FROM atividades WHERE tipo_atividade = 'Acadêmica Complementar'").fetchone()[0]
    total_atividades_extensao = conn.execute("SELECT COUNT(*) FROM atividades WHERE tipo_atividade = 'Extensão Universitária'").fetchone()[0]
    total_atividades = total_atividades_academicas + total_atividades_extensao
    
    # Estatísticas de requisições
    total_requisicoes = conn.execute("SELECT COUNT(*) FROM requisicoes").fetchone()[0]
    requisicoes_pendentes = conn.execute("SELECT COUNT(*) FROM requisicoes WHERE status = 'Pendente'").fetchone()[0]
    
    # Requisições por tipo de atividade
    requisicoes_academicas = conn.execute("""
        SELECT COUNT(*) FROM requisicoes r 
        JOIN atividades a ON r.atividade_id = a.id 
        WHERE a.tipo_atividade = 'Acadêmica Complementar'
    """).fetchone()[0]
    
    requisicoes_extensao = conn.execute("""
        SELECT COUNT(*) FROM requisicoes r 
        JOIN atividades a ON r.atividade_id = a.id 
        WHERE a.tipo_atividade = 'Extensão Universitária'
    """).fetchone()[0]
    
    return render_template("admin_dashboard.html", 
                           total_alunos=total_alunos, 
                           total_atividades=total_atividades,
                           total_atividades_academicas=total_atividades_academicas,
                           total_atividades_extensao=total_atividades_extensao,
                           total_requisicoes=total_requisicoes,
                           requisicoes_pendentes=requisicoes_pendentes,
                           requisicoes_academicas=requisicoes_academicas,
                           requisicoes_extensao=requisicoes_extensao)

@app.route("/admin/importar_requisicoes", methods=["GET", "POST"])
@admin_required
def admin_importar_requisicoes():
    if request.method == "POST":
        usar_padrao = "usar_arquivo_padrao" in request.form
        arquivo_selecionado = request.files.get("arquivo_excel")

        arquivo_path = None
        if usar_padrao:
            arquivo_path = os.path.join(app.config["UPLOAD_FOLDER"], "Acompanhamento de atividades complementares.xlsx")
            if not os.path.exists(arquivo_path):
                flash("Arquivo padrão não encontrado na pasta de uploads.", "error")
                return render_template("admin_importar_requisicoes.html")
        elif arquivo_selecionado and arquivo_selecionado.filename != "":
            filename = secure_filename(arquivo_selecionado.filename)
            arquivo_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            arquivo_selecionado.save(arquivo_path)
        else:
            flash("Nenhum arquivo selecionado.", "error")
            return render_template("admin_importar_requisicoes.html")

        # --- LÓGICA DE IMPORTAÇÃO COM OPENPYXL --- 
        try:
            logger.info(f"Iniciando importação do arquivo: {arquivo_path} usando openpyxl")
            workbook = openpyxl.load_workbook(arquivo_path, data_only=True) # data_only=True para obter valores de fórmulas
            
            if "Requisições" not in workbook.sheetnames:
                flash("Aba 'Requisições' não encontrada na planilha.", "error")
                logger.error(f"Aba 'Requisições' não encontrada em {arquivo_path}")
                return render_template("admin_importar_requisicoes.html")

            sheet = workbook["Requisições"]
            logger.info(f"Planilha 'Requisições' lida. Total de linhas: {sheet.max_row}")

            conn = get_db_connection()
            sucesso_count = 0
            erro_count = 0
            erros_detalhes = []

            atividades_map = {normalize_header(row["nome"]): row["id"] for row in conn.execute("SELECT id, nome FROM atividades").fetchall()}
            logger.info(f"Cache de atividades criado: {len(atividades_map)} atividades.")

            data_solicitacao_hoje = datetime.date.today().strftime("%Y-%m-%d")

            # Iterar pelas linhas começando da linha 3 (ignorando cabeçalho)
            for row_index in range(3, sheet.max_row + 1):
                try:
                    # Coluna F (índice 6): Nome da Atividade
                    # Coluna H (índice 8): Data do Evento
                    # Coluna G (índice 7): Horas (assumindo Solicitadas/Deferidas)
                    # Coluna I (índice 9): Status
                    
                    # Obter valores das células ou 'Indisponível'/'0.0' se vazias
                    nome_atividade_raw = sheet.cell(row=row_index, column=6).value
                    data_evento_raw = sheet.cell(row=row_index, column=8).value
                    horas_raw = sheet.cell(row=row_index, column=7).value
                    status_raw = sheet.cell(row=row_index, column=9).value

                    # Ignorar linhas onde dados essenciais são None/vazios
                    if nome_atividade_raw is None or data_evento_raw is None or horas_raw is None:
                        logger.warning(f"Linha {row_index}: Dados essenciais (Atividade, Data Evento, Horas) ausentes, pulando.")
                        continue
                    
                    # Ignorar linhas completamente vazias (heurística)
                    if all(sheet.cell(row=row_index, column=c).value is None for c in range(1, sheet.max_column + 1)):
                        logger.info(f"Linha {row_index}: Parece ser uma linha vazia, pulando.")
                        continue

                    # --- Processamento dos dados --- 
                    aluno_id = None # Não associar matrícula
                    
                    # Atividade ID
                    nome_atividade_norm = normalize_header(str(nome_atividade_raw))
                    atividade_id = atividades_map.get(nome_atividade_norm)
                    if not atividade_id:
                        logger.warning(f"Linha {row_index}: Atividade '{nome_atividade_raw}' (normalizada: '{nome_atividade_norm}') não encontrada no banco, pulando.")
                        erro_count += 1
                        erros_detalhes.append(f"Linha {row_index}: Atividade '{nome_atividade_raw}' não encontrada")
                        continue

                    # Data Evento
                    data_evento = "Indisponível"
                    if isinstance(data_evento_raw, datetime.datetime):
                        data_evento = data_evento_raw.strftime("%Y-%m-%d")
                    elif data_evento_raw is not None:
                        # Tentar converter se for string ou número (pode ser data serial do Excel)
                        try:
                            # Se for número, tenta converter de serial
                            if isinstance(data_evento_raw, (int, float)):
                                data_evento_dt = openpyxl.utils.datetime.from_excel(data_evento_raw)
                                data_evento = data_evento_dt.strftime("%Y-%m-%d")
                            else: # Tenta como string
                                # Adicionar mais formatos se necessário
                                data_evento_dt = datetime.datetime.strptime(str(data_evento_raw).split()[0], '%Y-%m-%d') 
                                data_evento = data_evento_dt.strftime("%Y-%m-%d")
                        except (ValueError, TypeError) as e_date:
                            logger.warning(f"Linha {row_index}: Formato de Data Evento inválido ('{data_evento_raw}'). Usando 'Indisponível'. Erro: {e_date}")
                    
                    # Horas Solicitadas / Deferidas
                    horas_solicitadas = 0.0
                    horas_deferidas = None
                    if horas_raw is not None:
                        try:
                            horas_solicitadas = float(horas_raw)
                        except (ValueError, TypeError) as e_horas:
                            logger.warning(f"Linha {row_index}: Valor de Horas inválido ('{horas_raw}'). Usando 0.0. Erro: {e_horas}")
                    else: # Se horas_raw for None
                         logger.warning(f"Linha {row_index}: Valor de Horas ausente. Usando 0.0.")

                    # Status
                    status = "Pendente" # Default inicial
                    if status_raw is not None:
                        status_norm = normalize_header(str(status_raw))
                        if status_norm == "deferido":
                            status = "Deferida"
                            horas_deferidas = horas_solicitadas
                        elif status_norm == "deferido parcialmente":
                            status = "Deferida Parcialmente"
                            horas_deferidas = 0.0 # Default para parcial, pode precisar de ajuste
                        elif status_norm == "indeferido":
                            status = "Indeferida"
                            horas_deferidas = 0.0
                    
                    # Inserir no banco de dados
                    conn.execute("""
                        INSERT INTO requisicoes 
                        (aluno_id, atividade_id, data_solicitacao, data_evento, horas_solicitadas, status, horas_deferidas, observacao)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (aluno_id, atividade_id, data_solicitacao_hoje, data_evento, horas_solicitadas, status, horas_deferidas, f"Importado da planilha linha {row_index}"))
                    sucesso_count += 1

                except Exception as e:
                    logger.error(f"Erro inesperado ao processar linha {row_index}: {e}")
                    traceback.print_exc()
                    erro_count += 1
                    erros_detalhes.append(f"Linha {row_index}: Erro inesperado - {e}")

            conn.commit()
            
            logger.info(f"Importação concluída com openpyxl. Sucesso: {sucesso_count}, Erros/Pulados: {erro_count}")

            flash(f"{sucesso_count} requisições importadas com sucesso.", "success")
            if erro_count > 0:
                flash(f"{erro_count} linhas não puderam ser importadas ou foram puladas. Verifique o log app.log para detalhes.", "warning")
                # Opcional: Mostrar detalhes dos erros no flash
                # flash('Detalhes dos erros: ' + '; '.join(erros_detalhes[:5]), 'warning') # Limita a 5 erros

            return redirect(url_for("admin_requisicoes"))

        except FileNotFoundError:
             logger.error(f"Arquivo não encontrado para importação: {arquivo_path}")
             flash(f"Erro: Arquivo não encontrado em {arquivo_path}", "error")
             return render_template("admin_importar_requisicoes.html")
        except Exception as e:
            logger.error(f"Erro GERAL durante a importação com openpyxl: {e}")
            traceback.print_exc()
            flash(f"Ocorreu um erro grave durante a importação: {e}", "error")
            return render_template("admin_importar_requisicoes.html")

    return render_template("admin_importar_requisicoes.html")

@app.route("/admin/requisicoes")
@admin_required
def admin_requisicoes():
    status_filtro = request.args.get('status', 'Todas')
    conn = get_db_connection()
    query = """
        SELECT r.*, a.nome as AlunoNome, a.matricula as AlunoMatricula, a.turma as TurmaNumero, act.nome as AtividadeNome
        FROM requisicoes r
        LEFT JOIN alunos a ON r.aluno_id = a.id
        JOIN atividades act ON r.atividade_id = act.id
    """
    params = []
    if status_filtro != 'Todas':
        query += " WHERE r.status = ?"
        params.append(status_filtro)
    query += " ORDER BY r.data_solicitacao DESC"
    
    requisicoes = conn.execute(query, params).fetchall()
    
    return render_template("admin_requisicoes.html", requisicoes=requisicoes, status_atual=status_filtro)

@app.route("/admin/requisicao/<int:req_id>")
@admin_required
def admin_detalhes_requisicao(req_id):
    conn = get_db_connection()
    requisicao = conn.execute("""
        SELECT r.*, u.nome as AdminNome, a.nome as AlunoNome, a.matricula as AlunoMatricula, act.nome as AtividadeNome
        FROM requisicoes r
        LEFT JOIN usuarios u ON r.admin_id = u.id
        LEFT JOIN alunos a ON r.aluno_id = a.id
        JOIN atividades act ON r.atividade_id = act.id
        WHERE r.id = ?
    """, (req_id,)).fetchone()
    if not requisicao:
        flash("Requisição não encontrada.", "error")
        return redirect(url_for("admin_requisicoes"))
    return render_template("admin_detalhes_requisicao.html", requisicao=requisicao)

@app.route("/admin/processar_requisicao/<int:req_id>", methods=["GET", "POST"])
@admin_required
def admin_processar_requisicao(req_id):
    conn = get_db_connection()
    requisicao = conn.execute("""
        SELECT r.*, a.nome as atividade_nome, a.tem_limitacao, a.tipo_limitacao, 
               a.limite_horas_total, a.limite_horas_semestral, al.nome as aluno_nome
        FROM requisicoes r
        JOIN atividades a ON r.atividade_id = a.id
        LEFT JOIN alunos al ON r.aluno_id = al.id
        WHERE r.id = ?
    """, (req_id,)).fetchone()
    
    if not requisicao:
        flash("Requisição não encontrada.", "error")
        return redirect(url_for("admin_requisicoes"))

    if request.method == "POST":
        status = request.form["status"]
        horas_deferidas = request.form.get("horas_deferidas")
        observacao = request.form.get("observacao")
        admin_id = session["user_id"]
        data_processamento = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if status == "Deferida Parcialmente" and not horas_deferidas:
            flash("Horas deferidas são obrigatórias para status 'Deferida Parcialmente'.", "error")
            return redirect(url_for("admin_processar_requisicao", req_id=req_id))
        
        # Validação de limites se a atividade tem limitação
        if status in ["Deferida", "Deferida Parcialmente"] and requisicao["tem_limitacao"]:
            horas_a_deferir = float(horas_deferidas) if status == "Deferida Parcialmente" else float(requisicao["horas_solicitadas"])
            
            # Calcular horas já deferidas para esta atividade específica
            if requisicao["tipo_limitacao"] == "total":
                # Verificar limite total
                horas_ja_deferidas = conn.execute("""
                    SELECT COALESCE(SUM(
                        CASE 
                            WHEN status = 'Deferida' THEN horas_solicitadas
                            WHEN status = 'Deferida Parcialmente' THEN horas_deferidas
                            ELSE 0
                        END
                    ), 0) as total
                    FROM requisicoes 
                    WHERE aluno_id = ? AND atividade_id = ? AND status IN ('Deferida', 'Deferida Parcialmente')
                """, (requisicao["aluno_id"], requisicao["atividade_id"])).fetchone()[0]
                
                if horas_ja_deferidas + horas_a_deferir > requisicao["limite_horas_total"]:
                    flash(f"Erro: O aluno já possui {horas_ja_deferidas}h nesta atividade. O limite total é {requisicao['limite_horas_total']}h. Máximo que pode ser deferido: {requisicao['limite_horas_total'] - horas_ja_deferidas}h.", "error")
                    return redirect(url_for("admin_processar_requisicao", req_id=req_id))
            
            elif requisicao["tipo_limitacao"] == "semestral":
                # Verificar limite semestral (considerando ano/semestre atual)
                ano_atual = datetime.datetime.now().year
                semestre_atual = 1 if datetime.datetime.now().month <= 6 else 2
                
                horas_ja_deferidas_semestre = conn.execute("""
                    SELECT COALESCE(SUM(
                        CASE 
                            WHEN status = 'Deferida' THEN horas_solicitadas
                            WHEN status = 'Deferida Parcialmente' THEN horas_deferidas
                            ELSE 0
                        END
                    ), 0) as total
                    FROM requisicoes 
                    WHERE aluno_id = ? AND atividade_id = ? AND status IN ('Deferida', 'Deferida Parcialmente')
                    AND strftime('%Y', data_evento) = ? 
                    AND (
                        (? = 1 AND strftime('%m', data_evento) BETWEEN '01' AND '06') OR
                        (? = 2 AND strftime('%m', data_evento) BETWEEN '07' AND '12')
                    )
                """, (requisicao["aluno_id"], requisicao["atividade_id"], str(ano_atual), semestre_atual, semestre_atual)).fetchone()[0]
                
                if horas_ja_deferidas_semestre + horas_a_deferir > requisicao["limite_horas_semestral"]:
                    flash(f"Erro: O aluno já possui {horas_ja_deferidas_semestre}h nesta atividade no semestre atual. O limite semestral é {requisicao['limite_horas_semestral']}h. Máximo que pode ser deferido: {requisicao['limite_horas_semestral'] - horas_ja_deferidas_semestre}h.", "error")
                    return redirect(url_for("admin_processar_requisicao", req_id=req_id))
        
        if status != "Deferida Parcialmente":
            horas_deferidas = None # Limpar se não for parcialmente deferida

        conn.execute("""
            UPDATE requisicoes SET status = ?, horas_deferidas = ?, observacao = ?, data_processamento = ?, admin_id = ?
            WHERE id = ?
        """, (status, horas_deferidas, observacao, data_processamento, admin_id, req_id))
        conn.commit()
        flash("Requisição processada com sucesso.", "success")
        return redirect(url_for("admin_requisicoes"))

    return render_template("admin_processar_requisicao.html", requisicao=requisicao)

@app.route("/admin/atividades")
@admin_required
def admin_atividades():
    tipo_filtro = request.args.get('tipo', 'Todas')
    conn = get_db_connection()
    
    if tipo_filtro == 'Todas':
        atividades = conn.execute("SELECT * FROM atividades ORDER BY tipo_atividade, grupo, nome").fetchall()
    else:
        atividades = conn.execute("SELECT * FROM atividades WHERE tipo_atividade = ? ORDER BY grupo, nome", (tipo_filtro,)).fetchall()
    
    return render_template("admin_atividades.html", atividades=atividades, tipo_atual=tipo_filtro)

# Rotas específicas para cada tipo de atividade
@app.route("/admin/atividades/academicas")
@admin_required
def admin_atividades_academicas():
    conn = get_db_connection()
    atividades = conn.execute("SELECT * FROM atividades WHERE tipo_atividade = 'Acadêmica Complementar' ORDER BY grupo, nome").fetchall()
    return render_template("admin_atividades_academicas.html", atividades=atividades)

@app.route("/admin/atividades/extensao")
@admin_required
def admin_atividades_extensao():
    conn = get_db_connection()
    atividades = conn.execute("SELECT * FROM atividades WHERE tipo_atividade = 'Extensão Universitária' ORDER BY grupo, nome").fetchall()
    return render_template("admin_atividades_extensao.html", atividades=atividades)

@app.route("/admin/adicionar_atividade", methods=["GET", "POST"])
@admin_required
def admin_adicionar_atividade():
    if request.method == "POST":
        grupo = request.form["grupo"]
        nome = request.form["nome"]
        limite_horas = request.form["limite_horas"]
        tipo_atividade = request.form["tipo_atividade"]
        
        # Novos campos de limitação
        tem_limitacao = 1 if request.form.get("tem_limitacao") else 0
        tipo_limitacao = request.form.get("tipo_limitacao") if tem_limitacao else None
        limite_horas_total = request.form.get("limite_horas_total") if tem_limitacao and tipo_limitacao == "total" else None
        limite_horas_semestral = request.form.get("limite_horas_semestral") if tem_limitacao and tipo_limitacao == "semestral" else None
        
        conn = get_db_connection()
        try:
            conn.execute("""
                INSERT INTO atividades 
                (grupo, nome, limite_horas, tipo_atividade, tem_limitacao, tipo_limitacao, limite_horas_total, limite_horas_semestral) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (grupo, nome, limite_horas, tipo_atividade, tem_limitacao, tipo_limitacao, limite_horas_total, limite_horas_semestral))
            conn.commit()
            flash("Atividade adicionada com sucesso.", "success")
            return redirect(url_for("admin_atividades"))
        except sqlite3.IntegrityError:
            flash("Erro: Atividade com este nome já existe.", "error")
    return render_template("admin_adicionar_atividade.html")

@app.route("/admin/editar_atividade/<int:atividade_id>", methods=["GET", "POST"])
@admin_required
def admin_editar_atividade(atividade_id):
    conn = get_db_connection()
    atividade = conn.execute("SELECT * FROM atividades WHERE id = ?", (atividade_id,)).fetchone()
    if not atividade:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("admin_atividades"))

    if request.method == "POST":
        grupo = request.form["grupo"]
        nome = request.form["nome"]
        limite_horas = request.form["limite_horas"]
        tipo_atividade = request.form["tipo_atividade"]
        
        # Novos campos de limitação
        tem_limitacao = 1 if request.form.get("tem_limitacao") else 0
        tipo_limitacao = request.form.get("tipo_limitacao") if tem_limitacao else None
        limite_horas_total = request.form.get("limite_horas_total") if tem_limitacao and tipo_limitacao == "total" else None
        limite_horas_semestral = request.form.get("limite_horas_semestral") if tem_limitacao and tipo_limitacao == "semestral" else None
        
        try:
            conn.execute("""
                UPDATE atividades 
                SET grupo = ?, nome = ?, limite_horas = ?, tipo_atividade = ?, 
                    tem_limitacao = ?, tipo_limitacao = ?, limite_horas_total = ?, limite_horas_semestral = ? 
                WHERE id = ?
            """, (grupo, nome, limite_horas, tipo_atividade, tem_limitacao, tipo_limitacao, limite_horas_total, limite_horas_semestral, atividade_id))
            conn.commit()
            flash("Atividade atualizada com sucesso.", "success")
            return redirect(url_for("admin_atividades"))
        except sqlite3.IntegrityError:
            flash("Erro: Atividade com este nome já existe.", "error")
    return render_template("admin_editar_atividade.html", atividade=atividade)

@app.route("/admin/deletar_atividade/<int:atividade_id>", methods=["POST"])
@admin_required
def admin_deletar_atividade(atividade_id):
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM atividades WHERE id = ?", (atividade_id,))
        conn.commit()
        flash("Atividade deletada com sucesso.", "success")
    except Exception as e:
        flash(f"Erro ao deletar atividade: {e}", "error")
    return redirect(url_for("admin_atividades"))

@app.route("/admin/alunos")
@admin_required
def admin_alunos():
    conn = get_db_connection()
    alunos = conn.execute("SELECT u.id as usuario_id, u.nome, u.email, a.matricula, a.turma, a.status FROM usuarios u JOIN alunos a ON u.id = a.usuario_id").fetchall()
    return render_template("admin_alunos.html", alunos=alunos)

@app.route("/admin/adicionar_aluno", methods=["GET", "POST"])
@admin_required
def admin_adicionar_aluno():
    if request.method == "POST":
        nome = request.form["nome"]
        email = request.form["email"]
        senha = request.form["senha"]
        matricula = request.form["matricula"]
        turma = request.form["turma"]
        status = request.form["status"]

        conn = get_db_connection()
        try:
            hashed_password = hash_password(senha)
            cursor = conn.execute("INSERT INTO usuarios (nome, email, senha, tipo) VALUES (?, ?, ?, ?)", (nome, email, hashed_password, "aluno"))
            usuario_id = cursor.lastrowid
            conn.execute("INSERT INTO alunos (usuario_id, nome, matricula, email, turma, status) VALUES (?, ?, ?, ?, ?, ?)", 
                         (usuario_id, nome, matricula, email, turma, status))
            conn.commit()
            flash("Aluno adicionado com sucesso.", "success")
            return redirect(url_for("admin_alunos"))
        except sqlite3.IntegrityError as e:
            if "UNIQUE constraint failed: usuarios.email" in str(e):
                flash("Erro: Já existe um usuário com este e-mail.", "error")
            elif "UNIQUE constraint failed: alunos.matricula" in str(e):
                flash("Erro: Já existe um aluno com esta matrícula.", "error")
            else:
                flash(f"Erro ao adicionar aluno: {e}", "error")
        except Exception as e:
            flash(f"Erro inesperado ao adicionar aluno: {e}", "error")

    return render_template("admin_adicionar_aluno.html")

@app.route("/admin/editar_aluno/<int:usuario_id>", methods=["GET", "POST"])
@admin_required
def admin_editar_aluno(usuario_id):
    conn = get_db_connection()
    aluno = conn.execute("SELECT u.id as usuario_id, u.nome, u.email, a.matricula, a.turma, a.status FROM usuarios u JOIN alunos a ON u.id = a.usuario_id WHERE u.id = ?", (usuario_id,)).fetchone()
    if not aluno:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("admin_alunos"))

    if request.method == "POST":
        nome = request.form["nome"]
        email = request.form["email"]
        matricula = request.form["matricula"]
        turma = request.form["turma"]
        status = request.form["status"]
        senha = request.form.get("senha") # Senha é opcional na edição

        try:
            # Atualizar tabela usuarios
            if senha:
                hashed_password = hash_password(senha)
                conn.execute("UPDATE usuarios SET nome = ?, email = ?, senha = ? WHERE id = ?", (nome, email, hashed_password, usuario_id))
            else:
                conn.execute("UPDATE usuarios SET nome = ?, email = ? WHERE id = ?", (nome, email, usuario_id))
            
            # Atualizar tabela alunos
            conn.execute("UPDATE alunos SET nome = ?, matricula = ?, email = ?, turma = ?, status = ? WHERE usuario_id = ?", 
                         (nome, matricula, email, turma, status, usuario_id))
            conn.commit()
            flash("Aluno atualizado com sucesso.", "success")
            return redirect(url_for("admin_alunos"))
        except sqlite3.IntegrityError as e:
            if "UNIQUE constraint failed: usuarios.email" in str(e):
                flash("Erro: Já existe outro usuário com este e-mail.", "error")
            elif "UNIQUE constraint failed: alunos.matricula" in str(e):
                flash("Erro: Já existe outro aluno com esta matrícula.", "error")
            else:
                flash(f"Erro ao atualizar aluno: {e}", "error")
        except Exception as e:
            flash(f"Erro inesperado ao atualizar aluno: {e}", "error")

    return render_template("admin_editar_aluno.html", aluno=aluno)

@app.route("/admin/deletar_aluno/<int:usuario_id>", methods=["POST"])
@admin_required
def admin_deletar_aluno(usuario_id):
    conn = get_db_connection()
    try:
        # Deletar primeiro da tabela alunos (que tem a FK para usuarios)
        conn.execute("DELETE FROM alunos WHERE usuario_id = ?", (usuario_id,))
        # Depois deletar da tabela usuarios
        conn.execute("DELETE FROM usuarios WHERE id = ?", (usuario_id,))
        conn.commit()
        flash("Aluno deletado com sucesso.", "success")
    except Exception as e:
        flash(f"Erro ao deletar aluno: {e}", "error")
    return redirect(url_for("admin_alunos"))

@app.route("/admin/alterar_status_alunos", methods=["POST"])
@admin_required
def admin_alterar_status_alunos():
    selected_alunos_ids = request.form.getlist("selected_alunos")
    novo_status = request.form["novo_status"]

    if not selected_alunos_ids:
        flash("Nenhum aluno selecionado para alteração de status.", "warning")
        return redirect(url_for("admin_alunos"))

    conn = get_db_connection()
    try:
        # Criar uma string de placeholders para a cláusula IN
        placeholders = ", ".join(["?" for _ in selected_alunos_ids])
        query = f"UPDATE alunos SET status = ? WHERE usuario_id IN ({placeholders})"
        
        # Os parâmetros devem ser uma tupla ou lista, com o novo_status primeiro
        conn.execute(query, (novo_status, *selected_alunos_ids))
        conn.commit()
        flash(f"Status de {len(selected_alunos_ids)} aluno(s) alterado(s) para '{novo_status}' com sucesso.", "success")
    except Exception as e:
        flash(f"Erro ao alterar status dos alunos: {e}", "error")
        logger.error(f"Erro ao alterar status em massa: {e}")
        traceback.print_exc()
    return redirect(url_for("admin_alunos"))


# --- Rotas Aluno ---

@app.route("/aluno/dashboard")
def aluno_dashboard():
    if "user_id" not in session or session["user_type"] != "aluno":
        flash("Acesso não autorizado.", "error")
        return redirect(url_for("login"))
    
    conn = get_db_connection()
    aluno_id = session["user_id"]

    # Obter informações do aluno
    aluno_info = conn.execute("SELECT * FROM alunos WHERE usuario_id = ?", (aluno_id,)).fetchone()
    if not aluno_info:
        flash("Dados do aluno não encontrados.", "error")
        return redirect(url_for("login"))

    # Obter requisições do aluno
    requisicoes = conn.execute("""
        SELECT r.*, act.nome as AtividadeNome, act.tipo_atividade
        FROM requisicoes r
        JOIN atividades act ON r.atividade_id = act.id
        WHERE r.aluno_id = ?
        ORDER BY r.data_solicitacao DESC
    """, (aluno_info["id"],)).fetchall()

    # Calcular horas totais deferidas por tipo de atividade
    horas_por_tipo = conn.execute("""
        SELECT act.tipo_atividade, SUM(r.horas_deferidas) as total_horas
        FROM requisicoes r
        JOIN atividades act ON r.atividade_id = act.id
        WHERE r.aluno_id = ? AND r.status IN ('Deferida', 'Deferida Parcialmente')
        GROUP BY act.tipo_atividade
    """, (aluno_info["id"],)).fetchall()

    # Calcular horas totais deferidas por grupo de atividade
    horas_por_grupo = conn.execute("""
        SELECT act.grupo, act.tipo_atividade, SUM(r.horas_deferidas) as total_horas
        FROM requisicoes r
        JOIN atividades act ON r.atividade_id = act.id
        WHERE r.aluno_id = ? AND r.status IN ('Deferida', 'Deferida Parcialmente')
        GROUP BY act.grupo, act.tipo_atividade
    """, (aluno_info["id"],)).fetchall()
    
    # Obter limites de horas por grupo
    limites_por_grupo = {row["grupo"]: row["limite_horas"] for row in conn.execute("SELECT grupo, limite_horas FROM atividades GROUP BY grupo").fetchall()}

    # Calcular total de horas deferidas por tipo
    total_horas_academicas = sum([r["horas_deferidas"] for r in requisicoes if r["status"] in ('Deferida', 'Deferida Parcialmente') and r["horas_deferidas"] is not None and r["tipo_atividade"] == "Acadêmica Complementar"])
    total_horas_extensao = sum([r["horas_deferidas"] for r in requisicoes if r["status"] in ('Deferida', 'Deferida Parcialmente') and r["horas_deferidas"] is not None and r["tipo_atividade"] == "Extensão Universitária"])

    return render_template("aluno_dashboard.html", 
                           aluno=aluno_info, 
                           requisicoes=requisicoes,
                           horas_por_tipo=horas_por_tipo,
                           horas_por_grupo=horas_por_grupo,
                           limites_por_grupo=limites_por_grupo,
                           total_horas_academicas=total_horas_academicas,
                           total_horas_extensao=total_horas_extensao)

@app.route("/aluno/nova_requisicao", methods=["GET", "POST"])
def aluno_nova_requisicao():
    if "user_id" not in session or session["user_type"] != "aluno":
        flash("Acesso não autorizado.", "error")
        return redirect(url_for("login"))

    conn = get_db_connection()
    tipo_filtro = request.args.get('tipo', 'Acadêmica Complementar')
    
    if tipo_filtro == 'Todas':
        atividades = conn.execute("SELECT * FROM atividades ORDER BY tipo_atividade, grupo, nome").fetchall()
    else:
        atividades = conn.execute("SELECT * FROM atividades WHERE tipo_atividade = ? ORDER BY grupo, nome", (tipo_filtro,)).fetchall()

    if request.method == "POST":
        aluno_usuario_id = session["user_id"]
        aluno_info = conn.execute("SELECT id FROM alunos WHERE usuario_id = ?", (aluno_usuario_id,)).fetchone()
        if not aluno_info:
            flash("Dados do aluno não encontrados.", "error")
            return redirect(url_for("login"))
        aluno_id = aluno_info["id"]

        atividade_id = request.form["atividade_id"]
        data_evento = request.form["data_evento"]
        horas_solicitadas = float(request.form["horas_solicitadas"])
        observacao = request.form.get("observacao")
        arquivo_comprovante = request.files.get("arquivo_comprovante")

        filename = None
        if arquivo_comprovante and arquivo_comprovante.filename != '':
            filename = secure_filename(arquivo_comprovante.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            arquivo_comprovante.save(filepath)
        
        data_solicitacao = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            conn.execute("""
                INSERT INTO requisicoes 
                (aluno_id, atividade_id, data_solicitacao, data_evento, horas_solicitadas, status, observacao, arquivo_comprovante)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (aluno_id, atividade_id, data_solicitacao, data_evento, horas_solicitadas, "Pendente", observacao, filename))
            conn.commit()
            flash("Requisição enviada com sucesso!", "success")
            return redirect(url_for("aluno_dashboard"))
        except Exception as e:
            flash(f"Erro ao enviar requisição: {e}", "error")
            logger.error(f"Erro ao enviar requisição: {e}")
            traceback.print_exc()

    return render_template("aluno_nova_requisicao.html", atividades=atividades, tipo_atual=tipo_filtro)

@app.route("/aluno/meus_dados", methods=["GET", "POST"])
def aluno_meus_dados():
    if "user_id" not in session or session["user_type"] != "aluno":
        flash("Acesso não autorizado.", "error")
        return redirect(url_for("login"))

    conn = get_db_connection()
    usuario_id = session["user_id"]
    aluno = conn.execute("SELECT u.nome, u.email, a.matricula, a.turma FROM usuarios u JOIN alunos a ON u.id = a.usuario_id WHERE u.id = ?", (usuario_id,)).fetchone()

    if request.method == "POST":
        nome = request.form["nome"]
        email = request.form["email"]
        matricula = request.form["matricula"]
        turma = request.form["turma"]
        senha = request.form.get("senha")

        try:
            if senha:
                hashed_password = hash_password(senha)
                conn.execute("UPDATE usuarios SET nome = ?, email = ?, senha = ? WHERE id = ?", (nome, email, hashed_password, usuario_id))
            else:
                conn.execute("UPDATE usuarios SET nome = ?, email = ? WHERE id = ?", (nome, email, usuario_id))
            
            conn.execute("UPDATE alunos SET nome = ?, matricula = ?, email = ?, turma = ? WHERE usuario_id = ?", 
                         (nome, matricula, email, turma, usuario_id))
            conn.commit()
            flash("Seus dados foram atualizados com sucesso!", "success")
            return redirect(url_for("aluno_dashboard"))
        except sqlite3.IntegrityError as e:
            if "UNIQUE constraint failed: usuarios.email" in str(e):
                flash("Erro: Já existe outro usuário com este e-mail.", "error")
            elif "UNIQUE constraint failed: alunos.matricula" in str(e):
                flash("Erro: Já existe outro aluno com esta matrícula.", "error")
            else:
                flash(f"Erro ao atualizar dados: {e}", "error")
        except Exception as e:
            flash(f"Erro inesperado ao atualizar dados: {e}", "error")

    return render_template("aluno_meus_dados.html", aluno=aluno)

@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

# --- Rotas de Autenticação ---

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        senha = request.form["senha"]
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM usuarios WHERE email = ?", (email,)).fetchone()

        if user and check_password(user["senha"], senha):
            session["user_id"] = user["id"]
            session["user_type"] = user["tipo"]
            session["user_name"] = user["nome"]
            flash("Login realizado com sucesso!", "success")
            if user["tipo"] == "admin":
                return redirect(url_for("admin_dashboard"))
            else:
                return redirect(url_for("aluno_dashboard"))
        else:
            flash("E-mail ou senha inválidos.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("user_id", None)
    session.pop("user_type", None)
    session.pop("user_name", None)
    flash("Você foi desconectado.", "info")
    return redirect(url_for("login"))

@app.route("/")
def index():
    return redirect(url_for("login"))




if __name__ == "__main__":
    with app.app_context():
        init_db()
    app.run(debug=False, host="0.0.0.0", port=5000)

